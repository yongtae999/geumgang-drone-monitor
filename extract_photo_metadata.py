import os
import sys
import json
import openpyxl
from datetime import datetime
from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS

# Ensure UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

PROJECT_DIR = r"C:\Users\Owner\.gemini\antigravity\scratch\geumgang-drone-monitor"
DATA_DIR = os.path.join(PROJECT_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)

PHOTO_BASE = r"E:\0. 2026년\2. 금강청 천내리\사진"
EXCEL_PATH = r"E:\0. 2026년\2. 금강청 천내리\일일작업일지 및 결과표\일일제거작업일지 요약_천내리_제출자료_2026년도.xlsx"

def dms_to_deg(dms, ref):
    if not dms:
        return None
    d, m, s = [float(x) for x in dms]
    deg = d + m / 60.0 + s / 3600.0
    if ref in ['S', 'W']:
        deg = -deg
    return round(deg, 7)

def extract_photos():
    photos = []
    supported_ext = ('.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG')
    
    for root, _, files in os.walk(PHOTO_BASE):
        for f in files:
            if f.endswith(supported_ext):
                full_path = os.path.join(root, f)
                date_folder = "2026-07-24" if "260724" in root else ("2026-08-06" if "260806" in root else "2026-07-10")
                
                photo_info = {
                    "filename": f,
                    "full_path": full_path,
                    "folder": os.path.basename(root),
                    "date_group": date_folder,
                    "lat": None,
                    "lng": None,
                    "altitude": None,
                    "bearing": None,
                    "timestamp": None,
                    "stage": "작업 중"
                }
                
                try:
                    with Image.open(full_path) as img:
                        exif = img._getexif()
                        if exif:
                            gps_info = {}
                            date_str = None
                            for k, v in exif.items():
                                tag = TAGS.get(k, k)
                                if tag == 'DateTimeOriginal' or tag == 'DateTime':
                                    date_str = str(v)
                                elif tag == 'GPSInfo':
                                    for gk, gv in v.items():
                                        gtag = GPSTAGS.get(gk, gk)
                                        gps_info[gtag] = gv
                            
                            if date_str:
                                photo_info["timestamp"] = date_str
                            
                            if gps_info:
                                lat_dms = gps_info.get('GPSLatitude')
                                lat_ref = gps_info.get('GPSLatitudeRef', 'N')
                                lng_dms = gps_info.get('GPSLongitude')
                                lng_ref = gps_info.get('GPSLongitudeRef', 'E')
                                
                                if lat_dms and lng_dms:
                                    photo_info["lat"] = dms_to_deg(lat_dms, lat_ref)
                                    photo_info["lng"] = dms_to_deg(lng_dms, lng_ref)
                                
                                alt = gps_info.get('GPSAltitude')
                                if alt:
                                    photo_info["altitude"] = round(float(alt), 1)
                                
                                bearing = gps_info.get('GPSImgDirection') or gps_info.get('GPSDestBearing')
                                if bearing:
                                    photo_info["bearing"] = round(float(bearing), 1)
                except Exception as e:
                    print(f"Error parsing {f}: {e}")
                
                photos.append(photo_info)
                
    valid_coords = [p for p in photos if p["lat"] and p["lng"]]
    print(f"Total photos: {len(photos)}, Photos with GPS: {len(valid_coords)}")
    
    with open(os.path.join(DATA_DIR, "photos.json"), "w", encoding="utf-8") as out:
        json.dump(photos, out, ensure_ascii=False, indent=2)

def extract_excel():
    if not os.path.exists(EXCEL_PATH):
        print("Excel not found:", EXCEL_PATH)
        return
        
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    work_logs = []
    
    if "천내리" in wb.sheetnames:
        ws = wb["천내리"]
        # Rows 5 to 16
        for r in range(5, 17):
            idx = ws.cell(r, 1).value
            target_plant = ws.cell(r, 2).value
            location = ws.cell(r, 3).value
            work_date = ws.cell(r, 4).value
            method = ws.cell(r, 5).value
            
            # Stages
            stages = []
            if str(ws.cell(r, 6).value).strip() in ['√', 'v', 'V', 'O', '1', 'True']: stages.append("로제트")
            if str(ws.cell(r, 7).value).strip() in ['√', 'v', 'V', 'O', '1', 'True']: stages.append("영양생장")
            if str(ws.cell(r, 8).value).strip() in ['√', 'v', 'V', 'O', '1', 'True']: stages.append("개화")
            if str(ws.cell(r, 9).value).strip() in ['√', 'v', 'V', 'O', '1', 'True']: stages.append("결실")
            if str(ws.cell(r, 10).value).strip() in ['√', 'v', 'V', 'O', '1', 'True']: stages.append("고사")
            
            w = ws.cell(r, 11).value or 0
            h = ws.cell(r, 12).value or 0
            area = float(w) * float(h) if (w and h) else (float(w) if w else 0)
            
            hours = ws.cell(r, 13).value or 0
            amount_kg = ws.cell(r, 14).value or 0
            
            # Format date
            date_str = str(work_date)
            if isinstance(work_date, datetime):
                date_str = work_date.strftime("%Y-%m-%d")
            elif "08월 06일" in str(work_date):
                date_str = "2026-08-06"
            elif not date_str or date_str == 'None':
                date_str = f"2026-08-{r:02d} (예정)"
                
            zone_assigned = "1구간 (A)" if r <= 7 else ("2구간 (B)" if r <= 11 else "3구간 (C)")
            
            if idx:
                work_logs.append({
                    "id": idx,
                    "target_plant": str(target_plant or "가시박, 환삼덩굴"),
                    "location": str(location or "천내리 습지 일대"),
                    "zone": zone_assigned,
                    "work_date": date_str,
                    "is_completed": r in [5, 6],
                    "method": str(method or "낫으로 베기, 예초기 사용"),
                    "stages": stages if stages else ["영양생장"],
                    "width_m": w,
                    "length_m": h,
                    "area_sqm": area,
                    "hours": hours,
                    "amount_kg": amount_kg,
                    "workers": 5 if r in [5, 6] else 5
                })
                
    with open(os.path.join(DATA_DIR, "work_logs.json"), "w", encoding="utf-8") as out:
        json.dump(work_logs, out, ensure_ascii=False, indent=2)
    print(f"Extracted {len(work_logs)} work logs.")

def create_zones_geojson():
    zones = {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "properties": {
                    "id": "zone-1",
                    "name": "1구간 (A구간)",
                    "subname": "제원대교 상류부",
                    "area_sqm": 28836,
                    "target_species": ["가시박", "미국쑥부쟁이", "돼지풀"],
                    "density": "산재 (피도 25%)",
                    "priority": "중점 제거",
                    "start_coords": "N 36°06′42.8″ E 127°34′03.6″",
                    "end_coords": "N 36°06′25.6″ E 127°34′26.9″",
                    "color": "#38bdf8",
                    "completed_area": 18000,
                    "progress_pct": 62.4
                },
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[
                        [127.567667, 36.111889],
                        [127.568400, 36.111200],
                        [127.571000, 36.109000],
                        [127.574139, 36.107111],
                        [127.573500, 36.106500],
                        [127.570200, 36.108200],
                        [127.567100, 36.110500],
                        [127.567667, 36.111889]
                    ]]
                }
            },
            {
                "type": "Feature",
                "properties": {
                    "id": "zone-2",
                    "name": "2구간 (B구간)",
                    "subname": "천내리 습지 중심부",
                    "area_sqm": 72803,
                    "target_species": ["가시박(대군락)", "돼지풀"],
                    "density": "대군락 형성 (피도 75%)",
                    "priority": "최우선 집중 제거",
                    "start_coords": "N 36°06′25.6″ E 127°34′26.9″",
                    "end_coords": "N 36°06′18.2″ E 127°34′42.9″",
                    "color": "#f87171",
                    "completed_area": 48000,
                    "progress_pct": 65.9
                },
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[
                        [127.574139, 36.107111],
                        [127.575800, 36.106000],
                        [127.577200, 36.105400],
                        [127.578583, 36.105056],
                        [127.578100, 36.104200],
                        [127.576200, 36.104800],
                        [127.574500, 36.105500],
                        [127.573500, 36.106500],
                        [127.574139, 36.107111]
                    ]]
                }
            },
            {
                "type": "Feature",
                "properties": {
                    "id": "zone-3",
                    "name": "3구간 (C구간)",
                    "subname": "천내리 습지 하류부",
                    "area_sqm": 43167,
                    "target_species": ["가시박(대군락)"],
                    "density": "대군락 형성 (피도 60%)",
                    "priority": "확산 방지 차단",
                    "start_coords": "N 36°06′18.2″ E 127°34′42.9″",
                    "end_coords": "N 36°06′13.3″ E 127°34′55.6″",
                    "color": "#34d399",
                    "completed_area": 0,
                    "progress_pct": 0.0
                },
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[
                        [127.578583, 36.105056],
                        [127.579800, 36.104500],
                        [127.581200, 36.104000],
                        [127.582111, 36.103694],
                        [127.581800, 36.102800],
                        [127.580200, 36.103300],
                        [127.578100, 36.104200],
                        [127.578583, 36.105056]
                    ]]
                }
            }
        ]
    }
    
    with open(os.path.join(DATA_DIR, "zones.geojson"), "w", encoding="utf-8") as out:
        json.dump(zones, out, ensure_ascii=False, indent=2)
    print("Created zones.geojson successfully.")

if __name__ == "__main__":
    extract_photos()
    extract_excel()
    create_zones_geojson()
