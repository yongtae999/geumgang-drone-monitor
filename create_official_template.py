import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_DIR = r"C:\Users\Owner\.gemini\antigravity\scratch\geumgang-drone-monitor"
TEMPLATE_PATH = os.path.join(PROJECT_DIR, "assets", "templates", "일일제거작업일지_표준템플릿.xlsx")

def create_official_standard_template():
    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # 1. Sheet 1: 일일작업결과표_표준양식 (공통 공식 서식)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "일일작업결과표_표준양식"
    ws1.views.sheetView[0].showGridLines = True

    # Styles
    font_title = Font(name="맑은 고딕", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="맑은 고딕", size=10, italic=True, color="64748B")
    font_th = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    font_sub_th = Font(name="맑은 고딕", size=10, bold=True, color="1E293B")
    font_data = Font(name="맑은 고딕", size=10, color="0F172A")
    font_sample = Font(name="맑은 고딕", size=10, italic=True, color="475569")
    
    fill_header_dark = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Navy
    fill_sub_th = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Slate light
    fill_sample = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_summary = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_th = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # Title Banner
    ws1.merge_cells("A1:N1")
    ws1["A1"] = "2026년 생태계교란생물 / 외래생물 제거사업 일일 작업결과표 (공통 표준양식)"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = align_left
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:N2")
    ws1["A2"] = "※ (사)야생생물관리협회 전국 지부 및 유역환경청 보고용 공통 표준 서식 (구간 구분 없이 사업 대상지 전체 단위로 작성)"
    ws1["A2"].font = font_subtitle
    ws1["A2"].alignment = align_left
    ws1.row_dimensions[2].height = 20

    # Headers (Row 4 & 5)
    headers_structure = [
        # (col_start, col_end, row_start, row_end, title)
        (1, 1, 4, 5, "구분\n(회차)"),
        (2, 2, 4, 5, "대상식물 / 생물\n(종명)"),
        (3, 3, 4, 5, "작업장소\n(대상지 일원)"),
        (4, 4, 4, 5, "작업일자\n(YYYY-MM-DD)"),
        (5, 5, 4, 5, "제거 / 포획방법"),
        (6, 10, 4, 4, "생육단계 (해당란 √ 표시)"),
        (11, 11, 4, 5, "제거면적\n(㎡)"),
        (12, 12, 4, 5, "소요시간\n(시간)"),
        (13, 13, 4, 5, "제거량 / 포획량\n(kg 또는 마리)"),
        (14, 14, 4, 5, "투입인력\n(명)"),
        (15, 15, 4, 5, "제거종\n(상세)"),
        (16, 16, 4, 5, "비고 / 작업세부사항\n(특이사항)")
    ]

    for cs, ce, rs, re, text in headers_structure:
        if cs == ce and rs == re:
            cell = ws1.cell(row=rs, column=cs, value=text)
        else:
            ws1.merge_cells(start_row=rs, start_column=cs, end_row=re, end_column=ce)
            cell = ws1.cell(row=rs, column=cs, value=text)
        
        cell.font = font_th
        cell.fill = fill_header_dark
        cell.alignment = align_center

    # Sub-headers for 생육단계 (Row 5, Col 6~10)
    sub_stages = ["로제트", "영양생장", "개화", "결실", "고사"]
    for idx, stage in enumerate(sub_stages, start=6):
        cell = ws1.cell(row=5, column=idx, value=stage)
        cell.font = font_th
        cell.fill = fill_header_dark
        cell.alignment = align_center

    # Apply borders to all header cells
    for r in range(4, 6):
        for c in range(1, 17):
            ws1.cell(row=r, column=c).border = border_th
            if not ws1.cell(row=r, column=c).fill.fill_type:
                ws1.cell(row=r, column=c).fill = fill_header_dark

    ws1.row_dimensions[4].height = 24
    ws1.row_dimensions[5].height = 22

    # Sample / Starter Rows (Row 6 ~ 10)
    sample_data = [
        (1, "가시박, 환삼덩굴", "천내리 습지 일대", "2026-07-24", "손으로 뿌리뽑기", "√", "√", "", "", "", 18000, 6, 80, 4, "가시박, 환삼덩굴", "발아기 유묘 및 로제트 손 뿌리뽑기 집중"),
        (2, "가시박, 환삼덩굴", "천내리 습지 일대", "2026-08-06", "낫으로 베기, 예초기 사용", "", "√", "", "", "", 48000, 6, 800, 5, "가시박, 환삼덩굴", "영양생장기 대군락지 예초 및 지상부 낫베기"),
        (3, "황소개구리, 미국수련", "두웅습지 일원", "2026-08-20", "포획통발 가동, 뿌리 굴취", "", "√", "√", "", "", 1200, 6, 350, 5, "미국수련, 황소개구리", "수생식물 지하경 굴취 및 황소개구리 포획통발 15개소 가동"),
        (4, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""),
        (5, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""),
        (6, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""),
        (7, "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""),
        (8, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "")
    ]

    for row_idx, row_values in enumerate(sample_data, start=6):
        ws1.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            
            # Alignments
            if col_idx in [1, 4, 6, 7, 8, 9, 10]:
                cell.alignment = align_center
            elif col_idx in [11, 12, 13, 14]:
                cell.alignment = align_right
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
            else:
                cell.alignment = align_left

            if row_idx <= 8 and val != "":
                cell.fill = fill_sample

    # Summary Row (Row 14)
    ws1.row_dimensions[14].height = 24
    ws1.merge_cells("A14:E14")
    sum_label = ws1.cell(row=14, column=1, value="누적 합계 (SUM)")
    sum_label.font = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
    sum_label.alignment = align_center
    sum_label.fill = fill_summary

    for c in range(1, 6):
        ws1.cell(row=14, column=c).border = border_cell
        ws1.cell(row=14, column=c).fill = fill_summary

    # Sum stages blank
    for c in range(6, 11):
        cell = ws1.cell(row=14, column=c, value="-")
        cell.alignment = align_center
        cell.font = Font(name="맑은 고딕", size=10, color="94A3B8")
        cell.border = border_cell
        cell.fill = fill_summary

    # Formula for Area Sum (Col 11)
    cell_area_sum = ws1.cell(row=14, column=11, value="=SUM(K6:K13)")
    cell_area_sum.font = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
    cell_area_sum.alignment = align_right
    cell_area_sum.border = border_cell
    cell_area_sum.fill = fill_summary
    cell_area_sum.number_format = "#,##0"

    # Formula for Hours Sum (Col 12)
    cell_hours_sum = ws1.cell(row=14, column=12, value="=SUM(L6:L13)")
    cell_hours_sum.font = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
    cell_hours_sum.alignment = align_right
    cell_hours_sum.border = border_cell
    cell_hours_sum.fill = fill_summary
    cell_hours_sum.number_format = "#,##0"

    # Formula for Kg Sum (Col 13)
    cell_kg_sum = ws1.cell(row=14, column=13, value="=SUM(M6:M13)")
    cell_kg_sum.font = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
    cell_kg_sum.alignment = align_right
    cell_kg_sum.border = border_cell
    cell_kg_sum.fill = fill_summary
    cell_kg_sum.number_format = "#,##0"

    # Formula for Workers Sum (Col 14)
    cell_workers_sum = ws1.cell(row=14, column=14, value="=SUM(N6:N13)")
    cell_workers_sum.font = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
    cell_workers_sum.alignment = align_right
    cell_workers_sum.border = border_cell
    cell_workers_sum.fill = fill_summary
    cell_workers_sum.number_format = "#,##0"

    # Blank remaining summary
    for c in [15, 16]:
        cell = ws1.cell(row=14, column=c, value="")
        cell.border = border_cell
        cell.fill = fill_summary

    # Column Widths
    col_widths = {
        'A': 8,   # 회차
        'B': 18,  # 대상식물
        'C': 20,  # 작업장소
        'D': 13,  # 작업일자
        'E': 24,  # 제거방법
        'F': 8,   # 로제트
        'G': 9,   # 영양생장
        'H': 8,   # 개화
        'I': 8,   # 결실
        'J': 8,   # 고사
        'K': 14,  # 제거면적
        'L': 10,  # 소요시간
        'M': 16,  # 제거량
        'N': 10,  # 투입인력
        'O': 18,  # 제거종
        'P': 30   # 비고
    }
    for col_letter, width in col_widths.items():
        ws1.column_dimensions[col_letter].width = width

    # -------------------------------------------------------------
    # 2. Sheet 2: 예산사용현황 (공통 산출내역 연동)
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="예산사용현황")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:F1")
    ws2["A1"] = "생태계교란생물 제거사업 예산 사용 및 집행 현황"
    ws2["A1"].font = font_title
    ws2["A1"].alignment = align_left
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:F2")
    ws2["A2"] = "(단위: 원, 부가세 포함 기준)"
    ws2["A2"].font = font_subtitle
    ws2["A2"].alignment = align_right
    ws2.row_dimensions[2].height = 18

    # Table Header (Row 4)
    th2 = ["구분", "예산 세부항목", "예산액 (원)", "집행액 (원)", "잔액 (원)", "집행률 (%)"]
    for col_idx, text in enumerate(th2, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=text)
        cell.font = font_th
        cell.fill = fill_header_dark
        cell.alignment = align_center
        cell.border = border_th
    ws2.row_dimensions[4].height = 24

    budget_data = [
        ("인건비", "특별인부 (작업원 인건비)", 10175490, "= 일일작업결과표_표준양식!N14 * 226122", "=C5-D5", "=D5/C5"),
        ("인건비", "관리자 인건비 (연구보조원 등)", 1993000, 0, "=C6-D6", "=D6/C6"),
        ("경비", "조사단 운영비 (책임/전문조사원)", 615124, 0, "=C7-D7", "=D7/C7"),
        ("경비", "회의비 및 사전 안전교육비", 200000, 100000, "=C8-D8", "=D8/C8"),
        ("경비", "재료비 (예초기 날, 소모품 등)", 450000, 260000, "=C9-D9", "=D9/C9"),
        ("경비", "홍보비 (현수막 제작 등)", 30000, 30000, "=C10-D10", "=D10/C10"),
        ("경비", "결과보고서 제작 및 인쇄비", 250000, 0, "=C11-D11", "=D11/C11"),
        ("경비", "안전보건관리비 (상해보험료 등)", 455000, 0, "=C12-D12", "=D12/C12"),
        ("일반관리비", "일반관리비", 831600, 0, "=C13-D13", "=D13/C13")
    ]

    for row_idx, row_vals in enumerate(budget_data, start=5):
        ws2.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_cell
            
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx == 2:
                cell.alignment = align_left
            elif col_idx in [3, 4, 5]:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            elif col_idx == 6:
                cell.alignment = align_right
                cell.number_format = "0.0%"

    # Total Sum Row (Row 14)
    ws2.row_dimensions[14].height = 24
    cell_tot_lbl = ws2.cell(row=14, column=1, value="총 계")
    cell_tot_lbl.font = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
    cell_tot_lbl.alignment = align_center
    cell_tot_lbl.fill = fill_summary
    cell_tot_lbl.border = border_cell

    ws2.cell(row=14, column=2, value="사업비 총괄 합계").font = Font(name="맑은 고딕", size=10, bold=True, color="1E3A8A")
    ws2.cell(row=14, column=2).alignment = align_left
    ws2.cell(row=14, column=2).fill = fill_summary
    ws2.cell(row=14, column=2).border = border_cell

    for c_idx, form in [(3, "=SUM(C5:C13)"), (4, "=SUM(D5:D13)"), (5, "=SUM(E5:E13)"), (6, "=D14/C14")]:
        c = ws2.cell(row=14, column=c_idx, value=form)
        c.font = Font(name="맑은 고딕", size=11, bold=True, color="1E3A8A")
        c.fill = fill_summary
        c.border = border_cell
        c.alignment = align_right
        c.number_format = "0.0%" if c_idx == 6 else "#,##0"

    # Col Widths for Sheet 2
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 16
    ws2.column_dimensions['F'].width = 14

    # Save
    os.makedirs(os.path.dirname(TEMPLATE_PATH), exist_ok=True)
    wb.save(TEMPLATE_PATH)
    print(f"Successfully generated official standard template at: {TEMPLATE_PATH}")

if __name__ == "__main__":
    create_official_standard_template()
