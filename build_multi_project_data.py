import os
import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

PROJECT_DIR = r"C:\Users\Owner\.gemini\antigravity\scratch\geumgang-drone-monitor"
DATA_DIR = os.path.join(PROJECT_DIR, "data")
TEMPLATES_DIR = os.path.join(PROJECT_DIR, "assets", "templates")
os.makedirs(TEMPLATES_DIR, exist_ok=True)

def create_standard_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "일일작업결과표_표준양식"
    
    title_font = Font(name="맑은 고딕", size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    
    hdr_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    sub_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    data_font = Font(name="맑은 고딕", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    
    thin = Side(border_style="thin", color="CBD5E1")
    box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Title
    ws.merge_cells("A1:P1")
    ws["A1"] = "생태계교란생물 / 외래생물 제거사업 일일 작업결과표 (표준양식)"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 35
    
    col_names = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P"]
    
    # Row 2 & 3 Values
    ws["A2"] = "회차\n(No)"
    ws["B2"] = "작업일자\n(YYYY-MM-DD)"
    ws["C2"] = "작업구간\n(구간명)"
    ws["D2"] = "대상생물\n(종명)"
    ws["E2"] = "작업장소\n(세부지명)"
    ws["F2"] = "제거/포획방법"
    ws["G2"] = "로제트"
    ws["H2"] = "영양생장"
    ws["I2"] = "개화/성체"
    ws["J2"] = "결실/유생"
    ws["K2"] = "고사/난괴"
    ws["L2"] = "제거면적\n(㎡)"
    ws["M2"] = "제거량/포획량\n(kg / 마리)"
    ws["N2"] = "투입인력\n(명)"
    ws["O2"] = "소요시간\n(시간)"
    ws["P2"] = "작업세부 및 현장 특이사항"
    
    for col in col_names:
        cell = ws[f"{col}2"]
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box_border
    ws.row_dimensions[2].height = 32

    # Sample rows
    samples = [
        [1, "2026-08-20", "1구간", "미국수련, 마름", "두웅습지 개방수면부", "수거망·지하경 굴취", "", "√", "√", "", "", 1200, 350, 5, 6, "지하경 중심 집중 굴취 및 육상 반출"],
        [2, "2026-08-25", "2구간", "황소개구리", "수변부 갈대군락", "포획통발·뜰채", "", "", "√", "√", "", 2500, 85, 4, 6, "포획통발 15개소 설치 및 성체 42마리 포획"],
        [3, "2026-09-02", "1구간", "미국수련", "목재데크 관찰로 주변", "수거망 사용", "", "√", "", "", "", 800, 200, 5, 6, "2차 신규 발아 개체 반복 수거"]
    ]
    
    for r_idx, s in enumerate(samples, start=3):
        for c_idx, val in enumerate(s):
            cell = ws[f"{col_names[c_idx]}{r_idx}"]
            cell.value = val
            cell.font = data_font
            cell.alignment = center if c_idx in [0,1,2,6,7,8,9,10,13,14] else (right if c_idx in [11,12] else left)
            cell.border = box_border
        ws.row_dimensions[r_idx].height = 22
            
    # Set Column widths
    widths = [8, 14, 14, 18, 22, 20, 9, 10, 11, 11, 11, 14, 15, 10, 10, 32]
    for c_idx, w in enumerate(widths):
        ws.column_dimensions[col_names[c_idx]].width = w

    template_path = os.path.join(TEMPLATES_DIR, "일일제거작업일지_표준템플릿.xlsx")
    wb.save(template_path)
    print(f"Created standard template: {template_path}")

def create_projects_data():
    projects = [
        {
            "id": "cheonnaeri",
            "name": "금강청 천내리습지 생태계교란식물 제거사업",
            "short_name": "천내리습지 (가시박 등)",
            "agency": "금강유역환경청",
            "contractor": "(사)야생생물관리협회 충남지부",
            "location_name": "충청남도 금산군 천내리습지 일원 (제원대교)",
            "center_coords": [127.5755, 36.1065],
            "zoom": 15.2,
            "pitch": 65,
            "bearing": 115,
            "target_species": "가시박, 환삼덩굴, 돼지풀 등",
            "total_area_sqm": 144806,
            "total_budget": 15000000,
            "target_kg": 18830,
            "target_workers": 45,
            "zones_file": "data/zones.geojson",
            "work_logs_file": "data/work_logs.json",
            "photos_file": "data/photos.json",
            "kpis": {
                "total_target_area": 144806,
                "cum_removed_area": 66000,
                "progress_pct": 45.6,
                "cum_removed_kg": 880,
                "target_kg": 18830,
                "cum_workers": 10,
                "target_workers": 45,
                "total_budget": 15000000,
                "spent_budget": 1294488,
                "budget_pct": 8.6
            },
            "flight_tour": [
                { "lng": 127.5676, "lat": 36.1118, "alt": 220, "pitch": 65, "bearing": 135, "speed": 12.0, "name": "1구간 (A) 제원대교 상류" },
                { "lng": 127.5710, "lat": 36.1090, "alt": 180, "pitch": 60, "bearing": 125, "speed": 14.5, "name": "1구간 (A) 완충 구역" },
                { "lng": 127.5741, "lat": 36.1071, "alt": 150, "pitch": 70, "bearing": 110, "speed": 10.0, "name": "2구간 (B) 천내리습지 진입부" },
                { "lng": 127.5765, "lat": 36.1055, "alt": 130, "pitch": 68, "bearing": 105, "speed": 8.5,  "name": "2구간 (B) 가시박 대군락지 중심" },
                { "lng": 127.5785, "lat": 36.1050, "alt": 140, "pitch": 65, "bearing": 95,  "speed": 9.0,  "name": "2-3구간 경계부" },
                { "lng": 127.5815, "lat": 36.1038, "alt": 160, "pitch": 62, "bearing": 90,  "speed": 11.0, "name": "3구간 (C) 습지 하류부" }
            ]
        },
        {
            "id": "doowoong",
            "name": "2026년 두웅습지 외래생물 실태조사 및 확산방지 용역",
            "short_name": "두웅습지 (미국수련·황소개구리)",
            "agency": "금강유역환경청 자연환경과",
            "contractor": "(사)야생생물관리협회 대전·충남·세종지부",
            "location_name": "충청남도 태안군 원북면 신두리 두웅습지보호지역 (람사르습지)",
            "center_coords": [126.1955, 36.8322],
            "zoom": 16.5,
            "pitch": 62,
            "bearing": 45,
            "target_species": "미국수련, 마름, 황소개구리 (금개구리 보호)",
            "total_area_sqm": 67050,
            "total_budget": 19500000,
            "target_kg": 12000,
            "target_workers": 35,
            "zones_file": "data/doowoong_zones.geojson",
            "work_logs_file": "data/doowoong_work_logs.json",
            "photos_file": "data/doowoong_photos.json",
            "kpis": {
                "total_target_area": 67050,
                "cum_removed_area": 24500,
                "progress_pct": 36.5,
                "cum_removed_kg": 3450,
                "target_kg": 12000,
                "cum_workers": 12,
                "target_workers": 35,
                "total_budget": 19500000,
                "spent_budget": 5240000,
                "budget_pct": 26.9
            },
            "flight_tour": [
                { "lng": 126.1940, "lat": 36.8308, "alt": 160, "pitch": 60, "bearing": 35, "speed": 9.0, "name": "두웅습지 남측 진입로 및 안내소 상공" },
                { "lng": 126.1952, "lat": 36.8318, "alt": 110, "pitch": 68, "bearing": 45, "speed": 7.5, "name": "1구간 개방수면부 미국수련 군락지" },
                { "lng": 126.1962, "lat": 36.8328, "alt": 100, "pitch": 70, "bearing": 55, "speed": 6.5, "name": "2구간 수변 갈대습지 및 황소개구리 포획통발 구역" },
                { "lng": 126.1970, "lat": 36.8335, "alt": 130, "pitch": 65, "bearing": 60, "speed": 8.0, "name": "3구간 신두리 해안사구 배후 완충지대 (금개구리 보호)" },
                { "lng": 126.1955, "lat": 36.8322, "alt": 200, "pitch": 55, "bearing": 20, "speed": 12.0, "name": "두웅습지보호지역 6.7만㎡ 전체 부감" }
            ]
        }
    ]

    with open(os.path.join(DATA_DIR, "projects.json"), "w", encoding="utf-8") as out:
        json.dump(projects, out, ensure_ascii=False, indent=2)
    print("Created projects.json successfully.")

def create_doowoong_geojson():
    zones = {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "properties": {
                    "id": "zone-1",
                    "name": "1구간 (수생식물 제거구역)",
                    "subname": "두웅습지 개방수면부",
                    "area_sqm": 18500,
                    "target_species": ["미국수련", "마름"],
                    "density": "수면 피도 65%",
                    "priority": "지하경 굴취 제거",
                    "color": "#38bdf8",
                    "completed_area": 12000,
                    "progress_pct": 64.9
                },
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[
                        [126.1948, 36.8315],
                        [126.1958, 36.8318],
                        [126.1963, 36.8326],
                        [126.1956, 36.8331],
                        [126.1947, 36.8325],
                        [126.1948, 36.8315]
                    ]]
                }
            },
            {
                "type": "Feature",
                "properties": {
                    "id": "zone-2",
                    "name": "2구간 (양서류 포획구역)",
                    "subname": "수변 갈대·마름 군락지",
                    "area_sqm": 26500,
                    "target_species": ["황소개구리(성체·유생·난괴)", "마름"],
                    "density": "포획통발 15개소 가동",
                    "priority": "성체·난괴 집중 수거",
                    "color": "#ef4444",
                    "completed_area": 12500,
                    "progress_pct": 47.2
                },
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[
                        [126.1942, 36.8310],
                        [126.1958, 36.8312],
                        [126.1972, 36.8324],
                        [126.1968, 36.8338],
                        [126.1950, 36.8336],
                        [126.1938, 36.8322],
                        [126.1942, 36.8310]
                    ]]
                }
            },
            {
                "type": "Feature",
                "properties": {
                    "id": "zone-3",
                    "name": "3구간 (보호종 완충구역)",
                    "subname": "사구 배후 및 금개구리 서식지",
                    "area_sqm": 22050,
                    "target_species": ["금개구리(Ⅱ급 보호)"],
                    "density": "혼획방지 정밀 모니터링",
                    "priority": "생태계 훼손 방지",
                    "color": "#10b981",
                    "completed_area": 0,
                    "progress_pct": 0.0
                },
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [[
                        [126.1935, 36.8305],
                        [126.1965, 36.8308],
                        [126.1982, 36.8328],
                        [126.1975, 36.8345],
                        [126.1945, 36.8342],
                        [126.1930, 36.8325],
                        [126.1935, 36.8305]
                    ]]
                }
            }
        ]
    }
    
    with open(os.path.join(DATA_DIR, "doowoong_zones.geojson"), "w", encoding="utf-8") as out:
        json.dump(zones, out, ensure_ascii=False, indent=2)
    print("Created doowoong_zones.geojson successfully.")

def create_doowoong_work_logs():
    logs = [
        {
            "id": 1,
            "target_plant": "미국수련 (지하경), 마름",
            "location": "충남 태안군 두웅습지 개방수면부",
            "zone": "1구간 (수생식물)",
            "work_date": "2026-08-12",
            "is_completed": True,
            "method": "수거망 사용, 지하경 굴취 육상반출",
            "stages": ["영양생장", "개화"],
            "area_sqm": 1200,
            "hours": 6,
            "amount_kg": 1800,
            "workers": 5
        },
        {
            "id": 2,
            "target_plant": "황소개구리 (성체 58마리, 유생)",
            "location": "두웅습지 수변 갈대습지 15개소",
            "zone": "2구간 (양서류)",
            "work_date": "2026-08-14",
            "is_completed": True,
            "method": "포획통발 15개 설치, 투망·뜰채",
            "stages": ["개화/성체", "결실/유생"],
            "area_sqm": 2500,
            "hours": 6,
            "amount_kg": 1650,
            "workers": 4
        },
        {
            "id": 3,
            "target_plant": "미국수련, 황소개구리",
            "location": "두웅습지 목재데크 탐방로 주변",
            "zone": "1구간 (수생식물)",
            "work_date": "2026-08-22 (예정)",
            "is_completed": False,
            "method": "수거망 및 뜰채",
            "stages": ["영양생장"],
            "area_sqm": 1500,
            "hours": 6,
            "amount_kg": 800,
            "workers": 5
        }
    ]
    with open(os.path.join(DATA_DIR, "doowoong_work_logs.json"), "w", encoding="utf-8") as out:
        json.dump(logs, out, ensure_ascii=False, indent=2)
    print("Created doowoong_work_logs.json successfully.")

def create_doowoong_photos():
    photos = [
        {
            "filename": "doowoong_waterlily_01.jpg",
            "rel_url": "assets/photos/P20260724_071637304_7032984A-8E2E-4A03-BE13-F779D2184C4A.JPG",
            "folder": "두웅습지260812",
            "date_group": "2026-08-12",
            "lat": 36.8320,
            "lng": 126.1953,
            "altitude": 24.5,
            "bearing": 45.0,
            "timestamp": "2026:08:12 08:30:15",
            "stage": "수생식물(미국수련) 제거"
        },
        {
            "filename": "doowoong_bullfrog_trap_02.jpg",
            "rel_url": "assets/photos/P20260806_070026068_9831C9C9-3CC3-435C-BFD7-3228A99122FC.JPG",
            "folder": "두웅습지260814",
            "date_group": "2026-08-14",
            "lat": 36.8326,
            "lng": 126.1965,
            "altitude": 26.0,
            "bearing": 65.0,
            "timestamp": "2026:08:14 09:15:20",
            "stage": "황소개구리 포획통발 설치"
        }
    ]
    with open(os.path.join(DATA_DIR, "doowoong_photos.json"), "w", encoding="utf-8") as out:
        json.dump(photos, out, ensure_ascii=False, indent=2)
    print("Created doowoong_photos.json successfully.")

if __name__ == "__main__":
    create_standard_excel_template()
    create_projects_data()
    create_doowoong_geojson()
    create_doowoong_work_logs()
    create_doowoong_photos()
